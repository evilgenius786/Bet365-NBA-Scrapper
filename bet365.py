import csv
import datetime
import json
import os
import time

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

t = 1
timeout = 5

debug = False

headless = True
images = False
max = False

incognito = True

headers = {
    "Player": "",
    "WH Points": "https://www.bet365.com/#/AC/B18/C20604387/D43/E181378/F43/",
    "WH Points Odds": "",
    "WH PAR": "https://www.bet365.com/#/AC/B18/C20604387/D43/E181390/F43/",
    "WH PAR Odds": "",
    "WH BS": "https://www.bet365.com/#/AC/B18/C20604387/D43/E181391/F43/",
    "WH BS Odds": "",
    "WH 3PT": "",
    "WH 3PT Odds": ""
}


def main():
    logo()
    driver = getChromeDriver()
    filename = f"WH NBA Betting Data {datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S')}.csv"
    print("Output file", filename)
    data = {}
    for key in headers.keys():
        if headers[key] != "":
            print(f"============{key}============")
            driver.get(headers[key])
            time.sleep(1)
            getElement(driver, '//div[@class="srb-ParticipantLabelWithTeam_Name "]')
            for div in driver.find_elements_by_xpath('//div[@class="src-FixtureSubGroup src-FixtureSubGroup_Closed "]'):
                try:
                    div.click()
                except:
                    pass
            teams = [div.text.strip() for div in
                     driver.find_elements_by_xpath('//div[@class="srb-ParticipantLabelWithTeam_Name "]')]
            scores = [div.text.strip().split('\n') for div in driver.find_elements_by_xpath(
                '//div[@class="gl-ParticipantCenteredStacked gl-Participant_General gl-Market_General-cn1 gl-ParticipantCenteredStacked-wide "]')]
            for i in range(len(teams)):
                if teams[i] not in data.keys():
                    data[teams[i]] = {}
                data[teams[i]][key] = scores[i][0]
                data[teams[i]][f"{key} Odds"] = scores[i][1]
                print(teams[i], scores[i])
    try:
        driver.close()
    except:
        pass
    try:
        driver.quit()
    except:
        pass
    rows = []
    row = {}
    for key in data.keys():
        row['Player'] = key
        for k in data[key].keys():
            row[k] = data[key][k]
        rows.append(row.copy())
    print(json.dumps(rows, indent=4))
    with open(filename.replace('csv', 'json'), "w", newline='') as f:
        json.dump(rows, f, indent=4)
    with open(filename, "w", newline='') as f:
        x = csv.DictWriter(f, fieldnames=headers.keys())
        x.writeheader()
        x.writerows(rows)
    print("Converting CSV to XSLX")
    cvrt(filename)
    print(f"Done!! Press any key. Output written to file {filename}")


def cvrt(filename):
    wb = Workbook()
    worksheet = wb.active
    with open(filename, 'r', encoding='utf8') as f:
        rows = [row for row in csv.reader(f)]
    column_widths = []
    for row in rows:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
            else:
                column_widths += [len(cell)]
    for i, column_width in enumerate(column_widths):
        worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width + 1

    for row in rows:
        worksheet.append(row)
    for col in worksheet.columns:
        for cell in col:
            cell.alignment = Alignment(horizontal='center')
    wb.save(filename.replace('.csv', '.xlsx'))


def click(driver, xpath, js=False):
    if js:
        driver.execute_script("arguments[0].click();", getElement(driver, xpath))
    else:
        WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, xpath))).click()


def getElement(driver, xpath):
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.XPATH, xpath)))


def sendkeys(driver, xpath, keys, js=False):
    if js:
        driver.execute_script(f"arguments[0].value='{keys}';", getElement(driver, xpath))
    else:
        getElement(driver, xpath).send_keys(keys)


def getChromeDriver(proxy=None):
    options = webdriver.ChromeOptions()
    if debug:
        # print("Connecting existing Chrome for debugging...")
        options.debugger_address = "127.0.0.1:9222"
    else:
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument("--disable-blink-features")
        options.add_argument("--disable-blink-features=AutomationControlled")
    if not images:
        # print("Turning off images to save bandwidth")
        options.add_argument("--blink-settings=imagesEnabled=false")
    if headless:
        # print("Going headless")
        options.add_argument("--headless")
        options.add_argument("--window-size=1920x1080")
    if max:
        # print("Maximizing Chrome ")
        options.add_argument("--start-maximized")
    if proxy:
        # print(f"Adding proxy: {proxy}")
        options.add_argument(f"--proxy-server={proxy}")
    if incognito:
        # print("Going incognito")
        options.add_argument("--incognito")
    return webdriver.Chrome(options=options)


def getFirefoxDriver():
    options = webdriver.FirefoxOptions()
    if not images:
        # print("Turning off images to save bandwidth")
        options.set_preference("permissions.default.image", 2)
    if incognito:
        # print("Enabling incognito mode")
        options.set_preference("browser.privatebrowsing.autostart", True)
    if headless:
        # print("Hiding Firefox")
        options.add_argument("--headless")
        options.add_argument("--window-size=1920x1080")
    return webdriver.Firefox(options)


def logo():
    os.system('color 0a')
    print(r"""
    __________          __  ________    ________ .________
    \______   \  ____ _/  |_\_____  \  /  _____/ |   ____/
     |    |  _/_/ __ \\   __\ _(__  < /   __  \  |____  \ 
     |    |   \\  ___/ |  |  /       \\  |__\  \ /       \
     |______  / \___  >|__| /______  / \_____  //______  /
            \/      \/             \/        \/        \/ 
==================================================================
              bet365.com/ betting odds scraper by:
                   github.com/evilgenius786
==================================================================
[+] Output in XLSX/CSV/JSON
__________________________________________________________________                                                     
""")


if __name__ == '__main__':
    main()
